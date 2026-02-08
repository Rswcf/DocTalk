from __future__ import annotations

import io
from typing import List

from .base import ExtractedPage


def extract_xlsx(file_bytes: bytes) -> List[ExtractedPage]:
    """Extract text from XLSX files. Each sheet = one page.

    Formats data as markdown tables for rich rendering in the frontend.
    """
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    pages: List[ExtractedPage] = []

    for i, sheet_name in enumerate(wb.sheetnames, start=1):
        ws = wb[sheet_name]
        rows: List[List[str]] = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(cell) if cell is not None else '' for cell in row]
            # Skip entirely empty rows
            if any(c.strip() for c in cells):
                rows.append(cells)

        if not rows:
            continue

        # Normalize column count to maximum
        max_cols = max(len(r) for r in rows)
        for r in rows:
            while len(r) < max_cols:
                r.append('')

        # Build markdown table
        lines: List[str] = []
        # Header row (first row)
        lines.append('| ' + ' | '.join(rows[0]) + ' |')
        # Separator
        lines.append('| ' + ' | '.join('---' for _ in range(max_cols)) + ' |')
        # Data rows
        for row in rows[1:]:
            lines.append('| ' + ' | '.join(row) + ' |')

        text = '\n'.join(lines)
        if text.strip():
            pages.append(ExtractedPage(
                page_number=i,
                text=text,
                section_title=sheet_name,
            ))

    wb.close()

    if not pages:
        pages.append(ExtractedPage(page_number=1, text='(empty spreadsheet)'))

    return pages
